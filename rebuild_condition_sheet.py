from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


COLORS = ["白色", "蓝色", "绿色", "红色", "紫色"]


def find_workbook() -> Path:
    root = Path(".")
    matches = [p for p in root.glob("*.xlsx") if "修改2" in p.name and ".pre" not in p.name]
    matches.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    if not matches:
        raise FileNotFoundError("No target workbook found")
    return matches[0]


def classify_condition_part(part: str) -> tuple[str, str] | None:
    text = str(part or "").strip()
    if not text:
        return None
    if text in {"若那样做的话", "如果那样做的话", "这么做的话", "若如此做", "如果如此做"}:
        return None

    if text.startswith("共鸣："):
        names = re.findall(r"“([^”]+)”", text)
        if names:
            return "共鸣特定机师", " / ".join(names)
        feats = re.findall(r"特征<([^>]+)>", text)
        if feats:
            return "共鸣特征机师", " / ".join(feats)

    if text.startswith("搭乘中") or text.startswith("搭乘时"):
        if "<" in text and "机师" in text:
            return "搭乘中/搭乘时·特征机师", text.split("·", 1)[1] if "·" in text else text
        if any(color in text for color in COLORS) and "机师" in text:
            return "搭乘中/搭乘时·颜色机师", text.split("·", 1)[1] if "·" in text else text
        if ("Lv." in text or "等级" in text) and "机师" in text:
            return "搭乘中/搭乘时·等级机师", text.split("·", 1)[1] if "·" in text else text
        return "搭乘中/搭乘时", text

    if text.startswith("共鸣中") or text.startswith("共鸣时"):
        return "共鸣中/共鸣时", text

    if "我方战斗区中存在" in text and "机师" in text:
        return "战斗区存在机师", text

    if "才能攻击" in text:
        return "攻击许可条件", text

    if "期间" in text:
        return "期间条件", text

    if text.startswith("若") or "我方战斗区中存在" in text:
        return "其他显式条件", text

    return "其他条件", text


def main():
    path = find_workbook()
    wb = load_workbook(path)
    summary = wb["P2-1单卡汇总"]

    groups: dict[tuple[str, str], dict] = defaultdict(lambda: {"cards": set(), "sample_card": None, "sample_raw": None})

    for r in range(2, summary.max_row + 1):
        raw = str(summary.cell(r, 7).value or "").strip()
        if not raw:
            continue
        card_id = summary.cell(r, 1).value
        card_name = summary.cell(r, 2).value
        for part in [p.strip() for p in raw.split("；") if p.strip()]:
            classified = classify_condition_part(part)
            if not classified:
                continue
            cond_class, cond_param = classified
            key = cond_class, cond_param
            groups[key]["cards"].add(card_id)
            if not groups[key]["sample_card"]:
                groups[key]["sample_card"] = card_name
                groups[key]["sample_raw"] = part

    if "条件整理表" in wb.sheetnames:
        index = wb.sheetnames.index("条件整理表")
        del wb["条件整理表"]
    else:
        index = len(wb.sheetnames)

    ws = wb.create_sheet("条件整理表", index)
    ws.append(["条件大类", "条件参数", "命中卡数", "示例卡", "条件原文示例"])

    for (cond_class, cond_param), data in sorted(groups.items(), key=lambda item: (item[0][0] or "", str(item[0][1] or ""))):
        ws.append([cond_class, cond_param, len(data["cards"]), data["sample_card"], data["sample_raw"]])

    wb.save(path)
    print(path)
    print(ws.max_row)


if __name__ == "__main__":
    main()
