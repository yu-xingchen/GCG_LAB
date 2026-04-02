import os
import re
from pathlib import Path

import openpyxl


def latest_target_file():
    files = [f for f in os.listdir(".") if f.lower().endswith(".xlsx") and "修改2" in f and ".pre" not in f]
    files.sort(key=lambda f: os.path.getmtime(f), reverse=True)
    if not files:
        raise FileNotFoundError("No target workbook found")
    return files[0]


def yaml_scalar(value):
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    text = str(value).replace("\r", "")
    if "\n" in text:
        lines = text.split("\n")
        return "|\n" + "\n".join(f"      {line}" for line in lines)
    if text == "":
        return '""'
    if re.fullmatch(r"[A-Za-z0-9_\-\.\u4e00-\u9fff（）/<>·：，；、「」“” ]+", text):
        return f'"{text}"'
    return '"' + text.replace("\\", "\\\\").replace('"', '\\"') + '"'


def main():
    path = latest_target_file()
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb["价值表"]

    out = []
    out.append("version: 1")
    out.append("source_workbook: " + yaml_scalar(path))
    out.append("sheet: " + yaml_scalar("价值表"))
    out.append("effects:")

    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if not name:
            continue
        row = {
            "status": ws.cell(r, 1).value,
            "name": name,
            "param1_name": ws.cell(r, 3).value,
            "param2_name": ws.cell(r, 4).value,
            "outputs": {
                "value": ws.cell(r, 5).value,
                "board_value": ws.cell(r, 6).value,
                "body_value": ws.cell(r, 7).value,
                "hand_plus": ws.cell(r, 8).value,
                "opponent_hand_plus": ws.cell(r, 9).value,
                "threat_power": ws.cell(r, 10).value,
                "threat_value": ws.cell(r, 11).value,
                "survival_power": ws.cell(r, 12).value,
                "survival_value": ws.cell(r, 13).value,
                "convert_power": ws.cell(r, 14).value,
                "convert_value": ws.cell(r, 15).value,
            },
            "notes": ws.cell(r, 16).value,
        }
        out.append(f'  - status: {yaml_scalar(row["status"])}')
        out.append(f'    name: {yaml_scalar(row["name"])}')
        out.append(f'    param1_name: {yaml_scalar(row["param1_name"])}')
        out.append(f'    param2_name: {yaml_scalar(row["param2_name"])}')
        out.append("    outputs:")
        for k, v in row["outputs"].items():
            out.append(f"      {k}: {yaml_scalar(v)}")
        notes = row["notes"]
        if isinstance(notes, str) and "\n" in notes:
            out.append("    notes: " + yaml_scalar(notes))
        else:
            out.append(f"    notes: {yaml_scalar(notes)}")

    output_path = Path("data") / "value_tables" / "value_table.yaml"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(out) + "\n", encoding="utf-8-sig")
    print(output_path)


if __name__ == "__main__":
    main()
