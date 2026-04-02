from __future__ import annotations

import re
from collections import OrderedDict, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook


DETAIL_HEADERS = [
    "编号",
    "卡名",
    "类型",
    "效果序号",
    "来源",
    "原始效果文本",
    "分类桶",
    "具体条件",
    "分类",
    "参数1",
    "参数1值",
    "参数2",
    "参数2值",
    "价值",
    "场值",
    "身材值",
    "手牌增加",
    "对方手牌增加",
    "威胁力",
    "威胁值",
    "生存力",
    "生存值",
    "转换力",
    "转换值",
    "估算说明",
]

SUMMARY_HEADERS = [
    "编号",
    "卡名",
    "类型",
    "行类型",
    "条件大类",
    "条件参数",
    "原始条件片段",
    "包含效果",
    "价值",
    "场值",
    "身材值",
    "手牌增加",
    "对方手牌增加",
    "威胁力",
    "威胁值",
    "生存力",
    "生存值",
    "转换力",
    "转换值",
    "备注",
]

CONDITION_HEADERS = ["条件大类", "条件参数", "命中卡数", "示例卡", "条件原文示例"]

METRIC_FIELDS = ["价值", "场值", "身材值", "手牌增加", "对方手牌增加", "威胁力", "威胁值", "生存力", "生存值", "转换力", "转换值"]

COLORS = ["白色", "蓝色", "绿色", "红色", "紫色"]


def find_workbook() -> Path:
    root = Path(".")
    matches = [p for p in root.glob("*.xlsx") if "修改2" in p.name and ".pre" not in p.name]
    matches.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    if not matches:
        raise FileNotFoundError("No target workbook found")
    return matches[0]


def normalize_resonance(card_type: str, raw) -> tuple[str, str, str] | None:
    if str(card_type or "").strip() != "机体":
        return None

    text = str(raw or "").strip()
    compact = re.sub(r"\s+", "", text)
    if not compact or compact in {"无", "-", "—", "——", "——————"}:
        return None
    if set(compact) <= {"-", "—", "/", "／"}:
        return None

    names = re.findall(r"“([^”]+)”", text)
    if names:
        return ("共鸣特定机师", " / ".join(names), f"共鸣：{text}")

    feats = re.findall(r"特征<([^>]+)>", text)
    if feats:
        return ("共鸣特征机师", " / ".join(feats), f"共鸣：{text}")

    return None


def to_number(value):
    if isinstance(value, (int, float)) or value is None:
        return value
    text = str(value).strip()
    if not text or "条件" in text:
        return value
    try:
        num = float(text)
    except ValueError:
        return value
    return int(num) if num.is_integer() else num


def parse_breakthrough_score(raw) -> float:
    try:
        x = float(raw)
    except Exception:
        return 0.0
    if x <= 0:
        return 0.0
    if x <= 2:
        return 1.0
    if x <= 4:
        return 1.25
    if x <= 5:
        return 1.5
    return 1.75


def build_card_map(card_sheet) -> dict[str, dict]:
    cards = {}
    for r in range(2, card_sheet.max_row + 1):
        cid = card_sheet.cell(r, 1).value
        if not cid:
            continue
        ctype = card_sheet.cell(r, 6).value
        resonance_raw = card_sheet.cell(r, 15).value
        cards[cid] = {
            "id": cid,
            "name": card_sheet.cell(r, 4).value,
            "type": ctype,
            "resonance": normalize_resonance(ctype, resonance_raw),
        }
    return cards


def build_value_map(value_sheet) -> dict[str, dict]:
    values = {}
    for r in range(2, value_sheet.max_row + 1):
        name = value_sheet.cell(r, 2).value
        if not name:
            continue
        values[str(name)] = {
            "价值": to_number(value_sheet.cell(r, 5).value),
            "场值": to_number(value_sheet.cell(r, 6).value),
            "身材值": to_number(value_sheet.cell(r, 7).value),
            "手牌增加": to_number(value_sheet.cell(r, 8).value),
            "对方手牌增加": to_number(value_sheet.cell(r, 9).value),
            "威胁力": to_number(value_sheet.cell(r, 10).value),
            "威胁值": to_number(value_sheet.cell(r, 11).value),
            "生存力": to_number(value_sheet.cell(r, 12).value),
            "生存值": to_number(value_sheet.cell(r, 13).value),
            "转换力": to_number(value_sheet.cell(r, 14).value),
            "转换值": to_number(value_sheet.cell(r, 15).value),
        }
    return values


def sheet_rows(ws) -> list[dict]:
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        rows.append({headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)})
    return rows


def split_sentences(text: str) -> list[str]:
    parts = re.split(r"[。；]", text)
    return [part.strip(" ，,：:") for part in parts if part.strip(" ，,：:")]


def clean_clause(text: str) -> str:
    clause = str(text or "").strip(" ，,：:")
    clause = re.sub(r"^(此后|然后|并且|且)", "", clause).strip(" ，,：:")
    return clause


def extract_tag_conditions(raw_text: str) -> list[tuple[str, str, str]]:
    conditions = []
    for tag in re.findall(r"【([^】]+)】", raw_text):
        tag = tag.strip()
        if tag in {"配置时", "爆发", "主要", "启动·主要", "启动·瞬动", "攻击时", "破坏时", "每回合1次"}:
            continue

        if tag.startswith("搭乘中") or tag.startswith("搭乘时"):
            if "·" in tag and "机师" in tag:
                detail = tag.split("·", 1)[1]
                if "<" in detail:
                    conditions.append(("搭乘中/搭乘时·特征机师", detail, tag))
                elif any(color in detail for color in COLORS):
                    conditions.append(("搭乘中/搭乘时·颜色机师", detail, tag))
                elif "Lv." in detail or "等级" in detail:
                    conditions.append(("搭乘中/搭乘时·等级机师", detail, tag))
                else:
                    conditions.append(("搭乘中/搭乘时", detail, tag))
            else:
                conditions.append(("搭乘中/搭乘时", tag, tag))
            continue

        if tag.startswith("共鸣中") or tag.startswith("共鸣时"):
            conditions.append(("共鸣中/共鸣时", tag, tag))

    return conditions


def extract_text_conditions(raw_text: str, effect_class: str) -> list[tuple[str, str, str]]:
    conditions = []
    plain = re.sub(r"【[^】]+】", "", raw_text)
    sentences = split_sentences(plain)

    for idx, sentence in enumerate(sentences):
        if "只有在" in sentence and "才能攻击" in sentence:
            part = sentence.split("只有在", 1)[1].split("才能攻击", 1)[0].strip()
            if "我方战斗区中存在" in part and "机师" in part:
                conditions.append(("战斗区存在机师", part, f"只有在{part}才能攻击"))
            else:
                conditions.append(("攻击许可条件", part, f"只有在{part}才能攻击"))
            continue

        if "期间" in sentence:
            part = sentence.split("期间", 1)[0].strip() + "期间"
            if "我方战斗区中存在" in part and "机师" in part:
                conditions.append(("战斗区存在机师", part, part))
            else:
                conditions.append(("期间条件", part, part))
            continue

        if sentence.startswith("若那样做的话") or sentence.startswith("如果那样做的话") or sentence.startswith("这么做的话"):
            if idx > 0:
                prev = clean_clause(sentences[idx - 1])
                if prev:
                    if "我方战斗区中存在" in prev and "机师" in prev:
                        conditions.append(("战斗区存在机师", prev, prev))
                    else:
                        conditions.append(("前置动作条件", prev, prev))
            later = re.search(r"(?:此后，)?若(.+?)(?:，则|，)", sentence)
            if later:
                part = clean_clause(later.group(1))
                if part and "那样做的话" not in part and "这么做的话" not in part:
                    conditions.append(("其他显式条件", part, f"若{part}"))
            continue

        if sentence.startswith("此后，若"):
            if effect_class == "抽牌":
                part = re.split(r"则|，", sentence[len("此后，若"):], maxsplit=1)[0]
                part = clean_clause(part)
                if part:
                    conditions.append(("其他显式条件", part, f"若{part}"))
            continue

        if sentence.startswith("若"):
            part = re.split(r"则|，", sentence[1:], maxsplit=1)[0]
            part = clean_clause(part)
            if part:
                if "我方战斗区中存在" in part and "机师" in part:
                    conditions.append(("战斗区存在机师", part, f"若{part}"))
                elif "我方战斗区中存在" in part:
                    conditions.append(("场面存在条件", part, f"若{part}"))
                else:
                    conditions.append(("其他显式条件", part, f"若{part}"))
            continue

        if "我方战斗区中存在" in sentence and "机师" in sentence:
            part = clean_clause(sentence)
            conditions.append(("战斗区存在机师", part, part))

    seen = set()
    dedup = []
    for item in conditions:
        if item not in seen:
            dedup.append(item)
            seen.add(item)
    return dedup


def infer_conditions(row: dict) -> list[tuple[str, str, str]]:
    raw_text = str(row.get("原始效果文本") or "")
    effect_class = str(row.get("分类") or "")
    conditions = extract_tag_conditions(raw_text) + extract_text_conditions(raw_text, effect_class)
    seen = set()
    dedup = []
    for item in conditions:
        if item not in seen:
            dedup.append(item)
            seen.add(item)
    return dedup


def join_condition_text(conditions: list[tuple[str, str, str]]) -> str | None:
    if not conditions:
        return None
    return "；".join(item[2] for item in conditions if item[2])


def cleanup_condition_text(row: dict):
    text = str(row.get("具体条件") or "").strip()
    if not text:
        return

    parts = [part.strip() for part in text.split("；") if part.strip()]
    cleaned = []
    for part in parts:
        if part in {"若那样做的话", "如果那样做的话", "这么做的话"}:
            continue
        cleaned.append(part)

    raw_text = str(row.get("原始效果文本") or "")
    if row.get("分类") == "抽牌":
        match = re.search(r"此后，若(.+?)(?:，则|，)", raw_text)
        if match:
            extra = f"若{clean_clause(match.group(1))}"
            if extra not in cleaned:
                cleaned.append(extra)

    row["具体条件"] = "；".join(cleaned) if cleaned else None


def add_inverse_outputs(row: dict, value_map: dict[str, dict], effect_name: str, note: str):
    source = value_map.get(effect_name) or {}
    for field in METRIC_FIELDS:
        value = source.get(field, 0)
        if isinstance(value, (int, float)):
            row[field] = -value
    row["估算说明"] = note


def make_hidden_cannot_attack_row(base_row: dict, value_map: dict[str, dict]) -> dict:
    row = dict(base_row)
    row["来源"] = "隐含"
    row["原始效果文本"] = "默认不能攻击（由“只有满足条件才能攻击”隐含）"
    row["分类桶"] = "无条件"
    row["具体条件"] = None
    row["分类"] = "不能踢人"
    row["参数1"] = "机体数量"
    row["参数1值"] = 1
    row["参数2"] = None
    row["参数2值"] = None
    source = value_map.get("不能踢人") or {}
    for field in METRIC_FIELDS:
        row[field] = source.get(field, 0)
    row["估算说明"] = "由“只有满足条件才能攻击”隐含出的默认不能攻击"
    return row


def fix_detail_rows(detail_rows: list[dict], card_map: dict[str, dict], value_map: dict[str, dict]) -> list[dict]:
    fixed = []
    for row in detail_rows:
        for field in METRIC_FIELDS:
            row[field] = to_number(row[field])

        card = card_map.get(row["编号"], {})
        classification = str(row["分类"] or "")
        source = str(row["来源"] or "")

        if classification == "共鸣":
            resonance = card.get("resonance")
            if source == "默认" and resonance:
                row["分类桶"] = "有条件"
                row["具体条件"] = resonance[2]
                row["估算说明"] = "机体真实共鸣，按价值表原值"
                fixed.append(row)
            continue

        if classification == "突破":
            score = parse_breakthrough_score(row["参数1值"])
            row["价值"] = score
            row["场值"] = score
            row["身材值"] = 0
            row["手牌增加"] = 0
            row["对方手牌增加"] = 0
            row["威胁力"] = 0
            row["威胁值"] = 1
            row["生存力"] = 0
            row["生存值"] = 0
            row["转换力"] = 0
            row["转换值"] = 0
            row["估算说明"] = f"突破{row['参数1值']}按分档结算，固定提供1威胁值"
            fixed.append(row)
            continue

        conditions = infer_conditions(row)
        condition_text = join_condition_text(conditions)

        if classification == "条件攻击限制" and "才能攻击" in str(row["原始效果文本"] or ""):
            fixed.append(make_hidden_cannot_attack_row(row, value_map))
            row["分类桶"] = "有条件"
            row["具体条件"] = condition_text
            add_inverse_outputs(row, value_map, "不能踢人", "满足攻击条件后，抵消默认的不能攻击")
            fixed.append(row)
            continue

        if row["分类桶"] in {"爆发", "爆发后有条件"}:
            if row["分类桶"] == "爆发后有条件" and condition_text:
                row["具体条件"] = condition_text
                cleanup_condition_text(row)
            fixed.append(row)
            continue

        if conditions:
            if card.get("type") == "机师" and any(item[0].startswith("搭乘中/搭乘时") for item in conditions):
                row["分类桶"] = "无条件"
                row["具体条件"] = None
            else:
                row["分类桶"] = "有条件"
                row["具体条件"] = condition_text
                cleanup_condition_text(row)

        fixed.append(row)

    return fixed


def classify_condition_part(part: str) -> tuple[str, str]:
    if part.startswith("共鸣："):
        names = re.findall(r"“([^”]+)”", part)
        if names:
            return ("共鸣特定机师", " / ".join(names))
        feats = re.findall(r"特征<([^>]+)>", part)
        if feats:
            return ("共鸣特征机师", " / ".join(feats))

    if part.startswith("搭乘中") or part.startswith("搭乘时"):
        if "<" in part and "机师" in part:
            return ("搭乘中/搭乘时·特征机师", part.split("·", 1)[1] if "·" in part else part)
        if any(color in part for color in COLORS) and "机师" in part:
            return ("搭乘中/搭乘时·颜色机师", part.split("·", 1)[1] if "·" in part else part)
        if ("Lv." in part or "等级" in part) and "机师" in part:
            return ("搭乘中/搭乘时·等级机师", part.split("·", 1)[1] if "·" in part else part)
        return ("搭乘中/搭乘时", part)

    if part.startswith("共鸣中") or part.startswith("共鸣时"):
        return ("共鸣中/共鸣时", part)

    if "我方战斗区中存在" in part and "机师" in part:
        return ("战斗区存在机师", part)

    if "才能攻击" in part:
        return ("攻击许可条件", part)

    if "期间" in part:
        return ("期间条件", part)

    if part.startswith("若") or "我方战斗区中存在" in part:
        return ("其他显式条件", part)

    return ("其他条件", part)


def normalize_condition_part(part: str) -> str | None:
    text = str(part or "").strip()
    if not text:
        return None
    if text in {"若那样做的话", "如果那样做的话", "这么做的话", "若如此做", "如果如此做"}:
        return None
    return text


def summarize_condition(bucket: str, condition_text):
    text = str(condition_text or "").strip()
    if bucket == "无条件":
        return ("无条件效果", "无条件", None, None, None)
    if bucket == "爆发":
        return ("爆发效果", "爆发", None, None, None)

    row_type = "条件效果" if bucket == "有条件" else "爆发后条件效果"
    if not text:
        return (row_type, "其他条件", None, None, None)

    parts = [part.strip() for part in text.split("；") if part.strip()]
    classes = []
    params = []
    raws = []
    for part in parts:
        cond_class, cond_param = classify_condition_part(part)
        classes.append(cond_class)
        params.append(cond_param)
        raws.append(part)

    note = "复合条件，需同时满足" if len(parts) > 1 else None
    class_text = " + ".join(dict.fromkeys(classes))
    param_text = "；".join(dict.fromkeys(str(p) for p in params if p))
    raw_text = "；".join(dict.fromkeys(raws))
    return (row_type, class_text, param_text or None, raw_text or None, note)


def rebuild_summary(detail_rows: list[dict]) -> list[dict]:
    groups = OrderedDict()
    for row in detail_rows:
        row_type, cond_class, cond_param, raw_cond, note = summarize_condition(row["分类桶"], row["具体条件"])
        key = (row["编号"], row["卡名"], row["类型"], row_type, cond_class, cond_param, raw_cond)
        if key not in groups:
            groups[key] = {
                "编号": row["编号"],
                "卡名": row["卡名"],
                "类型": row["类型"],
                "行类型": row_type,
                "条件大类": cond_class,
                "条件参数": cond_param,
                "原始条件片段": raw_cond,
                "effects": [],
                "备注": note,
            }
            for field in METRIC_FIELDS:
                groups[key][field] = 0

        if row["分类"] not in groups[key]["effects"]:
            groups[key]["effects"].append(row["分类"])
        for field in METRIC_FIELDS:
            value = row[field]
            if isinstance(value, (int, float)):
                groups[key][field] += value
        if not groups[key]["备注"] and note:
            groups[key]["备注"] = note

    summary_rows = []
    for group in groups.values():
        out = {key: group.get(key) for key in SUMMARY_HEADERS}
        out["包含效果"] = "；".join(group["effects"])
        summary_rows.append(out)
    return summary_rows


def rebuild_condition_sheet(summary_rows: list[dict]) -> list[dict]:
    groups: dict[tuple[str, str], dict] = defaultdict(lambda: {"cards": set(), "sample_card": None, "sample_raw": None})
    for row in summary_rows:
        if row["行类型"] not in {"条件效果", "爆发后条件效果"}:
            continue
        raw = str(row["原始条件片段"] or "").strip()
        if not raw:
            continue
        for part in [p.strip() for p in raw.split("；") if p.strip()]:
            part = normalize_condition_part(part)
            if not part:
                continue
            cond_class, cond_param = classify_condition_part(part)
            if not cond_class or cond_class in {"无条件", "爆发"}:
                continue
            key = (cond_class, cond_param)
            groups[key]["cards"].add(row["编号"])
            if not groups[key]["sample_card"]:
                groups[key]["sample_card"] = row["卡名"]
                groups[key]["sample_raw"] = part

    rows = []
    for (cond_class, cond_param), data in sorted(groups.items(), key=lambda item: (item[0][0] or "", str(item[0][1] or ""))):
        rows.append(
            {
                "条件大类": cond_class,
                "条件参数": cond_param,
                "命中卡数": len(data["cards"]),
                "示例卡": data["sample_card"],
                "条件原文示例": data["sample_raw"],
            }
        )
    return rows


def replace_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[dict]):
    if title in workbook.sheetnames:
        index = workbook.sheetnames.index(title)
        del workbook[title]
    else:
        index = len(workbook.sheetnames)
    ws = workbook.create_sheet(title, index)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])


def main():
    path = find_workbook()
    wb = load_workbook(path)

    card_sheet = wb["卡表"]
    value_sheet = wb["价值表"]
    detail_sheet = wb["P2-1单卡明细"]

    cards = build_card_map(card_sheet)
    value_map = build_value_map(value_sheet)
    detail_rows = sheet_rows(detail_sheet)

    fixed_detail_rows = fix_detail_rows(detail_rows, cards, value_map)
    summary_rows = rebuild_summary(fixed_detail_rows)
    condition_rows = rebuild_condition_sheet(summary_rows)

    replace_sheet(wb, "P2-1单卡明细", DETAIL_HEADERS, fixed_detail_rows)
    replace_sheet(wb, "P2-1单卡汇总", SUMMARY_HEADERS, summary_rows)
    replace_sheet(wb, "条件整理表", CONDITION_HEADERS, condition_rows)

    if "新P1效果表" in wb.sheetnames:
        del wb["新P1效果表"]

    wb.save(path)
    print(path)


if __name__ == "__main__":
    main()
