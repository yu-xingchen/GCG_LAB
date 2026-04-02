from __future__ import annotations

import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


COLORS = ["白色", "蓝色", "绿色", "红色", "紫色"]
FAKE_PARTS = {"若那样做的话", "如果那样做的话", "这么做的话", "若如此做", "如果如此做"}


def find_workbook() -> Path:
    root = Path(".")
    matches = [p for p in root.glob("*.xlsx") if "修改2" in p.name and ".pre" not in p.name]
    matches.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    if not matches:
        raise FileNotFoundError("No target workbook found")
    return matches[0]


def strip_effect_tail(text: str) -> str:
    s = str(text or "").strip()
    if not s:
        return s

    if s in FAKE_PARTS:
        return ""

    if s.startswith("只有在") and "才能攻击" in s:
        return s

    if s.startswith("共鸣："):
        return s

    if s.startswith("搭乘中") or s.startswith("搭乘时") or s.startswith("共鸣中") or s.startswith("共鸣时"):
        return s

    if "期间" in s:
        return s.split("期间", 1)[0].strip("，, ") + "期间"

    if "，则" in s:
        return s.split("，则", 1)[0].strip("，, ")

    if "。则" in s:
        return s.split("。则", 1)[0].strip("，, ")

    if "，并" in s and s.startswith("若"):
        return s.split("，并", 1)[0].strip("，, ")

    if "，且" in s and s.startswith("若"):
        return s.split("，且", 1)[0].strip("，, ")

    return s


def atomicize_condition_text(text: str) -> list[str]:
    s = str(text or "").strip()
    if not s:
        return []

    parts: list[str] = []

    if "，若" in s:
        left, right = s.split("，若", 1)
        left = left.strip("，, ")
        if left:
            parts.append(left)
        s = "若" + right

    if s.startswith("若") and "，则" in s:
        parts.append(s.split("，则", 1)[0].strip("，, "))
        return [p for p in parts if p and p not in FAKE_PARTS]

    if s.startswith("若") and "。则" in s:
        parts.append(s.split("。则", 1)[0].strip("，, "))
        return [p for p in parts if p and p not in FAKE_PARTS]

    if s.startswith("若") and "，并" in s:
        parts.append(s.split("，并", 1)[0].strip("，, "))
        return [p for p in parts if p and p not in FAKE_PARTS]

    if s.startswith("若") and "，且" in s:
        parts.append(s.split("，且", 1)[0].strip("，, "))
        return [p for p in parts if p and p not in FAKE_PARTS]

    stripped = strip_effect_tail(s)
    if stripped:
        parts.append(stripped)

    dedup: list[str] = []
    seen = set()
    for part in parts:
        if part and part not in seen and part not in FAKE_PARTS:
            dedup.append(part)
            seen.add(part)
    return dedup


def classify_atomic_condition(text: str) -> tuple[str, str] | None:
    s = strip_effect_tail(text)
    if not s:
        return None

    if s.startswith("共鸣："):
        names = re.findall(r"“([^”]+)”", s)
        if names:
            return "共鸣特定机师", " / ".join(names)
        feats = re.findall(r"特征<([^>]+)>", s)
        if feats:
            return "共鸣特征机师", " / ".join(feats)

    if s.startswith("搭乘中") or s.startswith("搭乘时"):
        if "<" in s and "机师" in s:
            return "搭乘中/搭乘时·特征机师", s.split("·", 1)[1] if "·" in s else s
        if any(color in s for color in COLORS) and "机师" in s:
            return "搭乘中/搭乘时·颜色机师", s.split("·", 1)[1] if "·" in s else s
        if ("Lv." in s or "等级" in s) and "机师" in s:
            return "搭乘中/搭乘时·等级机师", s.split("·", 1)[1] if "·" in s else s
        return "搭乘中/搭乘时", s

    if s.startswith("共鸣中") or s.startswith("共鸣时"):
        return "共鸣中/共鸣时", s

    if s.startswith("只有在") and "才能攻击" in s:
        middle = s.split("只有在", 1)[1].split("才能攻击", 1)[0].strip()
        return "攻击许可条件", middle

    if "我方战斗区中存在" in s and "机师" in s:
        return "战斗区存在机师", s

    if "期间" in s:
        return "期间条件", s

    if s.startswith("若"):
        return "其他显式条件", s

    if "我方战斗区中存在" in s:
        return "其他显式条件", s

    return "其他条件", s


def split_raw_conditions(raw: str) -> list[str]:
    parts = []
    for part in str(raw or "").split("；"):
        parts.extend(atomicize_condition_text(part.strip()))
    return parts


def rebuild_condition_sheet() -> tuple[Path, int]:
    path = find_workbook()
    wb = load_workbook(path)
    summary = wb["P2-1单卡汇总"]

    groups: dict[tuple[str, str], dict] = defaultdict(lambda: {"cards": set(), "sample_card": None, "sample_raw": None})

    for r in range(2, summary.max_row + 1):
        raw = summary.cell(r, 7).value
        if not raw:
            continue
        card_id = summary.cell(r, 1).value
        card_name = summary.cell(r, 2).value

        for part in split_raw_conditions(str(raw)):
            classified = classify_atomic_condition(part)
            if not classified:
                continue
            cond_class, cond_param = classified
            key = cond_class, cond_param
            groups[key]["cards"].add(card_id)
            if not groups[key]["sample_card"]:
                groups[key]["sample_card"] = card_name
                groups[key]["sample_raw"] = part

    if "条件整理表" in wb.sheetnames:
        index = wb.sheetnames.index("条件整理表")
        del wb["条件整理表"]
    else:
        index = len(wb.sheetnames)

    ws = wb.create_sheet("条件整理表", index)
    ws.append(["条件大类", "条件参数", "命中卡数", "示例卡", "条件原文示例"])

    for (cond_class, cond_param), data in sorted(groups.items(), key=lambda item: (item[0][0] or "", str(item[0][1] or ""))):
        ws.append([cond_class, cond_param, len(data["cards"]), data["sample_card"], data["sample_raw"]])

    wb.save(path)
    return path, ws.max_row


def export_condition_csv(workbook_path: Path) -> Path:
    out_dir = workbook_path.parent / "output" / "csv_snapshot"
    out_dir.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(workbook_path, read_only=True, data_only=False)
    ws = wb["条件整理表"]
    out_path = out_dir / "11_条件整理表.csv"
    with out_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(list(row))
    return out_path


def main():
    workbook_path, rows = rebuild_condition_sheet()
    csv_path = export_condition_csv(workbook_path)
    print(workbook_path)
    print(rows)
    print(csv_path)


if __name__ == "__main__":
    main()
