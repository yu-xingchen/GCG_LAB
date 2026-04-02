from __future__ import annotations

import csv
import re
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook


COLORS = ["白色", "蓝色", "绿色", "红色", "紫色", "黑色"]
FAKE_PARTS = {"若那样做的话", "如果那样做的话", "这么做的话", "若如此做", "如果如此做"}

HEADERS = [
    "编号",
    "卡名",
    "类型",
    "行类型",
    "条件大类",
    "条件参数",
    "原始条件片段",
    "包含效果",
    "价值",
    "场值",
    "身材值",
    "手牌增加",
    "对方手牌增加",
    "威胁力",
    "威胁值",
    "生存力",
    "生存值",
    "转换力",
    "转换值",
    "备注",
]

DETAIL_SHEET = "P2-1单卡明细"
SUMMARY_SHEET = "P2-1单卡汇总"


def find_workbook() -> Path:
    matches = [p for p in Path(".").glob("*.xlsx") if "修改2" in p.name and ".pre" not in p.name]
    matches.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    if not matches:
        raise FileNotFoundError("No target workbook found")
    return matches[0]


def normalize_condition_piece(text: str | None) -> str:
    s = str(text or "").strip()
    if not s or s in FAKE_PARTS:
        return ""

    if s.startswith("只有在") and "才能攻击" in s:
        middle = s.split("只有在", 1)[1].split("才能攻击", 1)[0].strip()
        return f"只有在{middle}才能攻击"

    if "期间" in s:
        return s.split("期间", 1)[0].strip("，。 ") + "期间"

    for marker in ("，则", "。则", "，并", "，且"):
        if s.startswith("若") and marker in s:
            return s.split(marker, 1)[0].strip("，。 ")

    return s


def split_atomic_conditions(text: str | None) -> list[str]:
    s = str(text or "").strip()
    if not s:
        return []

    parts: list[str] = []
    for raw in s.split("；"):
        raw = raw.strip()
        if not raw:
            continue
        for marker in ("，若", "。若"):
            if marker in raw:
                left, right = raw.split(marker, 1)
                left = normalize_condition_piece(left)
                if left:
                    parts.append(left)
                raw = "若" + right
                break
        cleaned = normalize_condition_piece(raw)
        if cleaned and cleaned not in FAKE_PARTS:
            parts.append(cleaned)

    deduped: list[str] = []
    seen: set[str] = set()
    for part in parts:
        if part not in seen:
            deduped.append(part)
            seen.add(part)
    return deduped


def classify_condition(text: str) -> tuple[str, str | None]:
    s = normalize_condition_piece(text)

    if s.startswith("共鸣："):
        names = re.findall(r"“([^”]+)”", s)
        if names:
            return "共鸣特定机师", " / ".join(names)
        traits = re.findall(r"特征<([^>]+)>", s)
        if traits:
            return "共鸣特征机师", " / ".join(traits)

    if s.startswith("搭乘中") or s.startswith("搭乘时"):
        if "<" in s and "机师" in s:
            return "搭乘中/搭乘时·特征机师", s.split("·", 1)[1] if "·" in s else s
        if any(color in s for color in COLORS) and "机师" in s:
            return "搭乘中/搭乘时·颜色机师", s.split("·", 1)[1] if "·" in s else s
        if ("Lv." in s or "等级" in s) and "机师" in s:
            return "搭乘中/搭乘时·等级机师", s.split("·", 1)[1] if "·" in s else s
        return "搭乘中/搭乘时", s

    if s.startswith("共鸣中") or s.startswith("共鸣时"):
        return "共鸣中/共鸣时", s

    if s.startswith("只有在") and "才能攻击" in s:
        middle = s.split("只有在", 1)[1].split("才能攻击", 1)[0].strip()
        return "攻击许可条件", middle

    if "我方战斗区中存在" in s and "机师" in s:
        return "战斗区存在机师", s

    if "期间" in s:
        return "期间条件", s

    if s.startswith("若"):
        return "其他显式条件", s

    if "我方战斗区中存在" in s:
        return "其他显式条件", s

    return "其他条件", s


def export_csv_from_sheet(ws, workbook_path: Path) -> Path:
    out_dir = workbook_path.parent / "output" / "csv_snapshot"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "10_P2-1单卡汇总.csv"
    with out_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(list(row))
    return out_path


def main() -> None:
    workbook_path = find_workbook()
    wb = load_workbook(workbook_path)
    detail = wb[DETAIL_SHEET]

    grouped: OrderedDict[tuple, list] = OrderedDict()

    for row in detail.iter_rows(min_row=2, values_only=True):
        card_id, card_name, card_type = row[0], row[1], row[2]
        bucket = row[6]
        raw_condition_text = row[7]
        effect_name = row[8]
        metrics = row[13:24]

        if bucket == "无条件":
            row_type = "无条件效果"
            cond_class = "无条件"
            cond_param = None
            raw_fragment = None
            note = None
        elif bucket == "爆发":
            row_type = "爆发效果"
            cond_class = "爆发"
            cond_param = None
            raw_fragment = None
            note = None
        else:
            row_type = "条件效果" if bucket == "有条件" else "爆发后条件效果"
            atomic_conditions = split_atomic_conditions(raw_condition_text)
            if not atomic_conditions:
                cond_class = "其他条件"
                cond_param = None
                raw_fragment = None
                note = None
            else:
                classified = [classify_condition(part) for part in atomic_conditions]
                cond_class = "；".join(item[0] for item in classified)
                cond_param = "；".join(item[1] or "" for item in classified) or None
                raw_fragment = "；".join(atomic_conditions)
                note = "复合条件，需同时满足" if len(atomic_conditions) > 1 else None

        key = (card_id, card_name, card_type, row_type, cond_class, cond_param, raw_fragment)
        if key not in grouped:
            grouped[key] = [
                card_id,
                card_name,
                card_type,
                row_type,
                cond_class,
                cond_param,
                raw_fragment,
                [],
                *([0] * 11),
                note,
            ]

        effects = grouped[key][7]
        if effect_name and effect_name not in effects:
            effects.append(effect_name)

        for i, value in enumerate(metrics, start=8):
            if isinstance(value, (int, float)):
                grouped[key][i] += value

        if note:
            grouped[key][19] = note

    if SUMMARY_SHEET in wb.sheetnames:
        index = wb.sheetnames.index(SUMMARY_SHEET)
        del wb[SUMMARY_SHEET]
    else:
        index = len(wb.sheetnames)

    ws = wb.create_sheet(SUMMARY_SHEET, index)
    ws.append(HEADERS)

    for item in grouped.values():
        item[7] = "；".join(item[7])
        ws.append(item)

    wb.save(workbook_path)
    csv_path = export_csv_from_sheet(ws, workbook_path)
    print(workbook_path)
    print(ws.max_row)
    print(csv_path)


if __name__ == "__main__":
    main()
