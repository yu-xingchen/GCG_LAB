import re
from pathlib import Path

import openpyxl
import yaml


def latest_target_file() -> Path:
    root = Path(".")
    files = [p for p in root.glob("*.xlsx") if "修改2" in p.name and ".pre" not in p.name]
    files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError("No target workbook found")
    return files[0]


def resonance_struct(card_type: str, text) -> dict:
    if str(card_type or "").strip() != "机体":
        return {"kind": None, "values": []}

    raw = str(text or "").strip()
    compact = re.sub(r"\s+", "", raw)
    if not compact or compact in {"无", "-", "—", "——", "——————"}:
        return {"kind": None, "values": []}
    if set(compact) <= {"-", "—", "/", "／"}:
        return {"kind": None, "values": []}

    names = re.findall(r"“([^”]+)”", raw)
    if names:
        return {"kind": "pilot_name", "values": names}

    feats = re.findall(r"特征<([^>]+)>", raw)
    if feats:
        return {"kind": "pilot_trait", "values": feats}

    return {"kind": None, "values": []}


def main():
    path = latest_target_file()
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb["卡表"]

    cards = []
    for r in range(2, ws.max_row + 1):
        cid = ws.cell(r, 1).value
        if not cid:
            continue

        card_type = ws.cell(r, 6).value
        card = {
            "id": cid,
            "pack": ws.cell(r, 2).value,
            "rarity": ws.cell(r, 3).value,
            "name": ws.cell(r, 4).value,
            "color": ws.cell(r, 5).value,
            "type": card_type,
            "lv": ws.cell(r, 7).value,
            "cost": ws.cell(r, 8).value,
            "ap": ws.cell(r, 9).value,
            "hp": ws.cell(r, 10).value,
            "terrain": {
                "space": ws.cell(r, 11).value == "○",
                "ground": ws.cell(r, 12).value == "○",
            },
            "text": ws.cell(r, 13).value or "",
            "traits": re.findall(r"<([^>]+)>", str(ws.cell(r, 14).value or "")),
            "resonance": resonance_struct(card_type, ws.cell(r, 15).value),
            "title_ref": ws.cell(r, 15).value,
            "series": ws.cell(r, 16).value,
        }
        cards.append(card)

    out = {
        "version": 1,
        "source_workbook": path.name,
        "sheet": "卡表",
        "cards": cards,
    }

    output_path = Path("data") / "cards" / "cards.yaml"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig") as file:
        yaml.safe_dump(out, file, allow_unicode=True, sort_keys=False)
    print(output_path)


if __name__ == "__main__":
    main()
